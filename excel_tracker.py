"""
excel_tracker.py
----------------
Appends new jobs to a running Excel tracker (jobs_tracker.xlsx).
Called automatically by run_job_search.py after each scrape run.

Columns:
    Date Found | Company | Title | Location | Source | URL | Status | Notes

'Status' and 'Notes' are blank — fill them in manually to track applications.
New jobs are appended; existing rows are never modified.
"""

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE_DIR      = Path(__file__).parent
TRACKER_FILE  = BASE_DIR / "jobs_tracker.xlsx"

HEADERS = ["Date Found", "Company", "Title", "Location", "Source", "URL", "Visa Sponsor?", "Status", "Notes"]
COL_WIDTHS = [14, 22, 45, 20, 18, 60, 14, 14, 25]

HEADER_FILL   = PatternFill("solid", fgColor="2C3E50")
HEADER_FONT   = Font(color="FFFFFF", bold=True)
NEW_FILL      = PatternFill("solid", fgColor="EAF4FB")   # light blue for new rows
SPONSOR_FILL  = PatternFill("solid", fgColor="D5F5E3")   # green for confirmed sponsors


def _setup_sheet(ws):
    """Apply headers and column widths to a fresh sheet."""
    ws.append(HEADERS)
    for i, (cell, width) in enumerate(zip(ws[1], COL_WIDTHS), 1):
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"


def append_jobs(jobs: list[dict], run_label: str = "Scheduled Run") -> int:
    """
    Append new jobs to jobs_tracker.xlsx.
    Returns the number of rows added.
    """
    if not jobs:
        return 0

    # Load or create workbook
    if TRACKER_FILE.exists():
        wb = load_workbook(TRACKER_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Job Tracker"
        _setup_sheet(ws)

    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

    for job in jobs:
        visa = job.get("visa_sponsorship")  # True / False / None
        visa_label = "✅ Yes" if visa is True else ""   # False jobs are already excluded

        row = [
            today,
            job.get("company", ""),
            job.get("title", ""),
            job.get("location", ""),
            job.get("source", ""),
            job.get("url", ""),
            visa_label,
            "",   # Status — fill manually
            "",   # Notes — fill manually
        ]
        ws.append(row)

        # Highlight confirmed sponsors in green; all other new rows in light blue
        row_num = ws.max_row
        fill = SPONSOR_FILL if visa is True else NEW_FILL
        for col in range(1, len(HEADERS) + 1):
            ws.cell(row=row_num, column=col).fill = fill

        # Make URL clickable
        url = job.get("url", "")
        if url:
            url_cell = ws.cell(row=row_num, column=6)
            url_cell.hyperlink = url
            url_cell.font = Font(color="2980B9", underline="single")
            url_cell.value = "Apply"

    wb.save(TRACKER_FILE)
    return len(jobs)
